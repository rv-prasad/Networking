#This code is used to read Excel File named devices.xlxs and read the IP address from Excel file stored in Coloumn_1
#It saves running and start-up config in seperate txt files.

import openpyxl
from netmiko import ConnectHandler
from getpass import getpass
import os

# --- Configuration ---
EXCEL_FILE = 'devices.xlsx'

def get_credentials():
    """Prompts the user for SSH username and password."""
    username = input("Enter your SSH username: ")
    password = getpass("Enter your SSH password: ")
    return username, passworda

def main():
    """
    Reads device IPs from Excel, runs commands, saves output,
    and updates the Excel file with the status using openpyxl.
    """
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: The file '{EXCEL_FILE}' was not found.")
        return

    username, password = get_credentials()

    try:
        # Load the workbook and select the active worksheet
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
    except Exception as e:
        print(f"Error loading the Excel file: {e}")
        return

    commands_to_run = ['show running-config', 'show startup-config']

    # Iterate through rows in the sheet, starting from row 1
    for row in range(1, sheet.max_row + 1):
        # Read the IP address from the first column (A)
        ip_address = sheet.cell(row=row, column=1).value
        
        # Skip empty rows
        if not ip_address:
            break
            
        ip = str(ip_address).strip()
        print(f"\n--- Connecting to {ip} ---")

        device = {
            'device_type': 'arista_eos',
            'ip':   ip,
            'username': username,
            'password': password,
        }
        
        try:
            with ConnectHandler(**device) as net_connect:
                # Find the hostname from the device prompt
                prompt = net_connect.find_prompt()
                hostname = prompt.rstrip('#>').strip() if prompt else ip
                print(f"Device hostname: {hostname}")

                # Execute commands and save output
                for command in commands_to_run:
                    output = net_connect.send_command(command)
                    filename = f"{hostname}_{command.replace(' ', '_')}.txt"
                    with open(filename, 'w') as f:
                        f.write(output)
                    print(f"Output saved to '{filename}'")
                
                print(f"Successfully processed {hostname} ({ip}).")
                # Write "Success" to the second column (B)
                sheet.cell(row=row, column=11, value="Success")

        except Exception as e:
            error_message = str(e).splitlines()[0]
            print(f"Failed to connect or run commands on {ip}: {error_message}")
            # Write the failure message to the second column (B)
            sheet.cell(row=row, column=11, value=f"Failed: {error_message}")
            
    # Save the changes back to the Excel file
    try:
        workbook.save(EXCEL_FILE)
        print(f"\nSuccessfully updated '{EXCEL_FILE}' with the status for each device.")
    except Exception as e:
        print(f"\nError: Could not save the Excel file. "
              f"Please ensure it is not open elsewhere. Error: {e}")

if __name__ == "__main__":
    main()
