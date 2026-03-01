#This code is used to login to a list of devices saved in Excel File named devices.xlsx and then run the set of commands.
#output of commands are savecd in txt flie with hostname of the files in the same folder as this code and devices.xlsx file.
import openpyxl 
import getpass
import sys
import telnetlib
import os
import xlrd
import xlwt 
from xlutils.copy import copy
import netmiko
from netmiko import ConnectHandler
import datetime
from datetime import date

today = date.today()

def check_ping(str):
    hostname = str
    response = os.system("ping -n 1 " + hostname)
    # and then check the response...
    if response == 0:
        pingstatus = "Active"
		
    else:
        pingstatus = "Ping failed"

    return pingstatus



# Give the location of the file 
path = "devices.xlsx"

# To open the workbook 
# workbook object is created

wb = openpyxl.load_workbook(path) 

#Get workbook active sheet object 
# from the active attribute
sheet = wb.active 

# Getting the value of maximum rows
# and column
max_row = sheet.max_row
max_column = sheet.max_column

for i in range (2, max_row+1) :
	device_hostname=sheet.cell(i,1).value
	status=check_ping(device_hostname)
	if (status == "Active") :
		try : 
			"""send_command is pattern-based. Meaning it searches for the device prompt to detect the end of output. for BaseConnection in netmiko, the time for each command to complete is 100 seconds but for Cisco devices is only 10 seconds because fast_cli default value is set to True. fast_cli simply multiplies the 100 seconds by 0.1 (100*0.1 = 10) just to make it faster, yet faster is not always the best option.

			You need to set fast_cli to False to disable the timeout."""
			
			cisco_881 = { 'device_type': 'cisco_ios', 'ip': '10.148.232.84', 'username': 'XXXXXX','password': 'XXXXXXX',"fast_cli": False, 'secret':'password'}
			cisco_881['username'] = input("Enter your SSH username: ")
			cisco_881['password'] = getpass("Enter your SSH password: ")
			cisco_881['secret'] = getpass("Enter your Enable password: ")
			cisco_881['ip']= device_hostname
			net_connect = ConnectHandler(**cisco_881)
			print(net_connect.find_prompt())
			net_connect.enable()
			print(net_connect.find_prompt())
			net_connect.send_command("terminal length 0")
			write=net_connect.send_command("write", expect_string=r"[OK]")
			print(write)
			running_config=net_connect.send_command("show running-config")
			startup_config=net_connect.send_command("show startup-config")
			cdp_neighbor=net_connect.send_command("show cdp nei")
			lldp_neighbor=net_connect.send_command("show lldp nei")
			interface_status=net_connect.send_command("show ip int brief")
			inventory=net_connect.send_command("show inventory")
			bgp_sum=net_connect.send_command("show ip bgp sum")
			eigrp_sum=net_connect.send_command("show ip eigrp nei")

			filename=device_hostname+"-"+str(today)+".txt"
			with open (filename, 'w') as file :
				file.writelines("\n SAVING CONFIG\n-------------------------------\n"+write+
					"\n\nRUNNING CONFIG\n-------------------------------\n\n"+running_config+
					"\n\n STARTUP CONFIG\n-------------------------------\n\n"+startup_config+
					"\n\n CDP NEIGHBOR\n-------------------------------\n\n"+cdp_neighbor+
					"\n\n LLDP NEIGHBOR\n-------------------------------\n\n"+lldp_neighbor+
					"\n\n INTERFACE STATUS\n-------------------------------n\n"+interface_status+
					"\n\nINVENTORY\n-------------------------------\n"+inventory+
					"\n\n BGP SUMMARY\n-------------------------------\n\n"+bgp_sum+
					"\n\n EIGRP NEIGHBOR\n-------------------------------\n\n"+eigrp_sum)

			
			
			sheet.cell(i,3).value="completed"
			net_connect.disconnect()
			i=i+1


		except:
			continue

	else :
		sheet.cell(i,3).value="Device Unreachable"

wb.save('backup_report.xlsx')
