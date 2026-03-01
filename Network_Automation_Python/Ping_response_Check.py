import openpyxl 
import sys
import telnetlib
import os
import subprocess
from pythonping import ping

# Give the location of the file 
path = "devices.xlsx"

# To open the workbook 
# workbook object is created

wb = openpyxl.load_workbook(path) 

#Get workbook active sheet object 
# from the active attribute
sheet = wb.active 

# Getting the value of maximum rows
# and column
# in Openpyxl the row and coloumns numbers starts with  number 1 like A1 will be Cell(1,1)
max_row = sheet.max_row
max_column = sheet.max_column

for i in range (2, max_row) :
	device_ip=sheet.cell(i,2).value
	print(device_ip)
	result=str(ping(device_ip, count=1, verbose=True))
	print(result)
	sheet.cell(i,7).value=str(result)

wb.save('devices_ping_result.xlsx')

