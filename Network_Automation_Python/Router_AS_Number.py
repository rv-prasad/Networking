# This code is used to login to devices if they are pinging and then get the AS number of BGP and then save it to the Excel file.
#Useful in case if you want to list All the AS number configured on devices.
import openpyxl 
import getpass
import sys
import telnetlib
import os
import xlrd
import xlwt 
from xlutils.copy import copy
import netmiko
from netmiko import ConnectHandler

def check_ping(str):
    hostname = str
    response = os.system("ping -n 1 " + hostname)
    # and then check the response...
    if response == 0:
        pingstatus = "Active"
		
    else:
        pingstatus = "Ping failed"

    return pingstatus



# Give the location of the file 
path = "Router_List2.xlsx"

# To open the workbook 
# workbook object is created

wb = openpyxl.load_workbook(path) 

#Get workbook active sheet object 
# from the active attribute
sheet = wb.active 

# Getting the value of maximum rows
# and column
max_row = sheet.max_row
max_column = sheet.max_column

for i in range (2, max_row+1) :
	device_ip=sheet.cell(i,2).value
	status=check_ping(device_ip)
	if (status == "Active") :
		try : 
			cisco_881 = { 'device_type': 'cisco_ios', 'ip': '10.148.232.84', 'username': 'XXXXXX','password': 'XXXXXXX','secret':'password'}
			cisco_881['username'] = input("Enter your SSH username: ")
			cisco_881['password'] = getpass("Enter your SSH password: ")
			cisco_881['ip']= device_ip
			net_connect = ConnectHandler(**cisco_881)
			AS_Output=net_connect.send_command("show ip bgp summary | in local AS number")
			print(AS_Output)
			output_Split=AS_Output.split()
			AS=output_Split[-1]
			sheet.cell(i,3).value=AS

		except:
			continue

	else :
		sheet.cell(i,3).value="Device Unreachable"

wb.save('AS Number List.xlsx')



