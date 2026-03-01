import openpyxl 
import getpass
import sys
import telnetlib
import os
import xlrd
import xlwt 
from xlutils.copy import copy
import netmiko
from netmiko import ConnectHandler
import datetime
from datetime import date

today = date.today()

def check_ping(str):
    hostname = str
    response = os.system("ping -n 1 " + hostname)
    # and then check the response...
    if response == 0:
        pingstatus = "Active"
		
    else:
        pingstatus = "Ping failed"

    return pingstatus



# Give the location of the file 
path = "devices.xlsx"

# To open the workbook 
# workbook object is created

wb = openpyxl.load_workbook(path) 
sheets = wb.sheetnames
print(sheets)

#Get sheetnames based upon the name of sheet printed on Above sheetname list
#Devices are saved in Excel worksheet named Routers
sheet = wb["Routers"]

# Getting the value of maximum rows
# and column
max_row = sheet.max_row
max_column = sheet.max_column


for i in range (2, max_row+1) :
	device_ip=sheet.cell(i,3).value
	print(device_ip)
	status=check_ping(device_ip)

	if (status == "Active") :
		try : 
			"""
			send_command is pattern-based. Meaning it searches for the device prompt to detect the end of output. for BaseConnection in netmiko, the time for each command to complete is 100 seconds but for Cisco devices is only 10 seconds because fast_cli default value is set to True. fast_cli simply multiplies the 100 seconds by 0.1 (100*0.1 = 10) just to make it faster, yet faster is not always the best option.

			You need to set fast_cli to False to disable the timeout.
			"""
			
			cisco_881 = { 'device_type': 'cisco_ios', 'ip': '10.148.232.84', 'username': 'XXXXXX','password': 'XXXXXXX',"fast_cli": False, 'secret':'password'}
			cisco_881['username'] = input("Enter your SSH username: ")
			cisco_881['password'] = getpass("Enter your SSH password: ")
			cisco_881['secret']=getpass("Enter your Enable password: ")
			cisco_881['ip']= device_ip
			net_connect = ConnectHandler(**cisco_881)
			print(net_connect.find_prompt())
			througput_command=net_connect.send_command("show ver | in throughput")
			print(througput_command)
			througput_command_split=througput_command.split()
			througput=througput_command_split[5]
			sheet.cell(i,5).value=througput
			net_connect.disconnect()
			i=i+1


		except:
			continue

	else :
		sheet.cell(i,12).value="Device Unreachable"

wb.save('Routers_with_throughput.xlsx')


